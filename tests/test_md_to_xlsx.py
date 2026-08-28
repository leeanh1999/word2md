"""Tests for the Markdown -> Excel direction."""

from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import load_workbook  # noqa: E402
from PIL import Image  # noqa: E402

from src.converter import (  # noqa: E402
    STATUS_SUCCESS,
    ConversionOptions,
    convert_file,
    output_suffix,
)
import src.md_to_xlsx as md_to_xlsx  # noqa: E402
from src.md_to_xlsx import (  # noqa: E402
    MAX_CELL_LENGTH,
    XlsxSettings,
    markdown_to_xlsx,
)


class XlsxTestCase(unittest.TestCase):
    """Writes a .md file, converts it, and hands back the workbook."""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.dir = Path(self.tmp.name)

    def tearDown(self):
        self.tmp.cleanup()

    def build(self, markdown: str, name: str = "tài liệu.md", **kwargs):
        source = self.dir / name
        source.write_text(markdown, encoding="utf-8")
        destination = self.dir / f"{source.stem}.xlsx"
        warnings = markdown_to_xlsx(source, destination, **kwargs)
        return load_workbook(destination), warnings

    def values(self, sheet) -> list[tuple]:
        return list(sheet.iter_rows(values_only=True))


class SheetSplitTests(XlsxTestCase):
    def test_title_then_sections_become_one_sheet_each(self):
        book, _ = self.build(
            "# Báo cáo\n\n## Doanh thu\n\nA\n\n## Chi phí\n\nB\n"
        )
        self.assertEqual(book.sheetnames, ["Doanh thu", "Chi phí"])

    def test_content_under_a_lone_title_keeps_its_own_sheet(self):
        book, _ = self.build("# Báo cáo\n\nMở đầu.\n\n## Doanh thu\n\nA\n")
        self.assertEqual(book.sheetnames, ["Báo cáo", "Doanh thu"])
        self.assertEqual(self.values(book["Báo cáo"]), [("Mở đầu.",)])

    def test_repeated_top_level_headings_split_the_sheets(self):
        book, _ = self.build("# Một\n\nA\n\n# Hai\n\nB\n")
        self.assertEqual(book.sheetnames, ["Một", "Hai"])

    def test_a_document_without_headings_is_named_after_the_file(self):
        book, _ = self.build("Chỉ có một đoạn.\n", name="ghi chú.md")
        self.assertEqual(book.sheetnames, ["ghi chú"])

    def test_no_title_option_falls_back_to_the_default_name(self):
        book, _ = self.build("Đoạn.\n", add_title_heading=False)
        self.assertEqual(book.sheetnames, ["Sheet1"])

    def test_an_empty_document_still_produces_a_workbook(self):
        book, warnings = self.build("\n\n", name="rỗng.md")
        self.assertEqual(book.sheetnames, ["rỗng"])
        self.assertEqual(self.values(book["rỗng"]), [])
        self.assertTrue(any("không có nội dung" in w for w in warnings), warnings)


class SheetNameTests(XlsxTestCase):
    def test_characters_excel_rejects_are_dropped(self):
        book, _ = self.build("# A[1]:B/C\n\nx\n\n# Khác\n\ny\n")
        self.assertEqual(book.sheetnames[0], "A1BC")

    def test_long_headings_are_cut_to_31_characters(self):
        title = "Đây là một tiêu đề rất dài vượt quá giới hạn của Excel"
        book, _ = self.build(f"# {title}\n\nx\n\n# Khác\n\ny\n")
        self.assertEqual(book.sheetnames[0], title[:31])

    def test_duplicate_headings_are_numbered(self):
        book, _ = self.build("# Bảng\n\nx\n\n# Bảng\n\ny\n\n# Bảng\n\nz\n")
        self.assertEqual(book.sheetnames, ["Bảng", "Bảng (2)", "Bảng (3)"])

    def test_inline_markup_is_stripped_from_the_name(self):
        book, _ = self.build("# **Doanh** thu\n\nx\n\n# Khác\n\ny\n")
        self.assertEqual(book.sheetnames[0], "Doanh thu")


class TableTests(XlsxTestCase):
    MARKDOWN = (
        "## Doanh thu\n\n"
        "| Mã | Số lượng | Ghi chú |\n"
        "|---|---:|:---:|\n"
        "| 007 | 1200 | **tốt** |\n"
        "| A1 | 3.5 | |\n"
    )

    def setUp(self):
        super().setUp()
        self.book, self.warnings = self.build(self.MARKDOWN)
        self.sheet = self.book["Doanh thu"]

    def test_cells_land_in_a_real_grid(self):
        self.assertEqual(
            self.values(self.sheet),
            [
                ("Mã", "Số lượng", "Ghi chú"),
                ("007", 1200, "tốt"),
                ("A1", 3.5, None),
            ],
        )

    def test_plain_numbers_become_numbers_and_the_rest_stays_text(self):
        self.assertIsInstance(self.sheet["B2"].value, int)
        self.assertIsInstance(self.sheet["B3"].value, float)
        self.assertEqual(self.sheet["A2"].value, "007", "a leading zero is data")

    def test_header_row_is_bold_filled_and_frozen(self):
        self.assertTrue(self.sheet["A1"].font.bold)
        self.assertEqual(self.sheet["A1"].fill.fgColor.rgb[-6:], "DDEBF7")
        self.assertEqual(self.sheet.freeze_panes, "A2")

    def test_column_alignment_follows_the_markdown(self):
        self.assertIsNone(self.sheet["A2"].alignment.horizontal)
        self.assertEqual(self.sheet["B2"].alignment.horizontal, "right")
        self.assertEqual(self.sheet["C2"].alignment.horizontal, "center")

    def test_columns_are_wide_enough_to_read(self):
        widths = {name: dim.width for name, dim in self.sheet.column_dimensions.items()}
        self.assertEqual(set(widths), {"A", "B", "C"})
        self.assertTrue(all(8 <= width <= 60 for width in widths.values()), widths)

    def test_a_ragged_row_is_padded_not_dropped(self):
        book, _ = self.build(
            "## T\n\n| A | B |\n|---|---|\n| chỉ một |\n", name="ragged.md"
        )
        self.assertEqual(self.values(book["T"]), [("A", "B"), ("chỉ một", None)])

    def test_a_table_after_text_does_not_freeze_the_pane(self):
        book, _ = self.build(
            "## T\n\nMở đầu.\n\n| A |\n|---|\n| 1 |\n", name="after.md"
        )
        self.assertIsNone(book["T"].freeze_panes)


class LinkAndImageTests(XlsxTestCase):
    def test_a_cell_that_is_only_a_link_becomes_a_hyperlink(self):
        book, _ = self.build(
            "## T\n\n| Nguồn |\n|---|\n| [Trang chủ](https://x.dev) |\n"
        )
        cell = book["T"]["A2"]
        self.assertEqual(cell.value, "Trang chủ")
        self.assertEqual(cell.hyperlink.target, "https://x.dev")
        self.assertEqual(cell.font.underline, "single")

    def test_a_link_inside_a_sentence_keeps_its_url_in_brackets(self):
        book, _ = self.build("## T\n\nXem [tại đây](https://y.dev) nhé.\n")
        cell = book["T"]["A1"]
        self.assertEqual(cell.value, "Xem tại đây (https://y.dev) nhé.")
        self.assertIsNone(cell.hyperlink)

    def test_an_autolink_is_not_repeated(self):
        book, _ = self.build("## T\n\n<https://z.dev>\n")
        self.assertEqual(book["T"]["A1"].value, "https://z.dev")


class BlockTests(XlsxTestCase):
    def test_headings_below_the_sheet_level_become_bold_rows(self):
        book, _ = self.build("# T\n\n## A\n\n### A.1\n\nnội dung\n")
        sheet = book["A"]
        self.assertEqual(self.values(sheet), [("A.1",), ("nội dung",)])
        self.assertTrue(sheet["A1"].font.bold)
        self.assertFalse(sheet["A2"].font.bold)

    def test_lists_keep_their_markers_and_nesting(self):
        book, _ = self.build(
            "## L\n\n- một\n- hai\n  - hai a\n\n1. đầu\n2. sau\n"
        )
        self.assertEqual(
            self.values(book["L"]),
            [("• một",), ("• hai",), ("◦ hai a",), ("1. đầu",), ("2. sau",)],
        )

    def test_task_items_show_their_state(self):
        book, _ = self.build("## L\n\n- [x] xong\n- [ ] chưa\n")
        self.assertEqual(
            self.values(book["L"]), [("• [x] xong",), ("• [ ] chưa",)]
        )

    def test_a_table_inside_a_list_item_is_still_a_grid(self):
        book, _ = self.build(
            "## L\n\n- mục\n\n    | A |\n    |---|\n    | 1 |\n"
        )
        self.assertIn(("A",), self.values(book["L"]))

    def test_code_keeps_a_monospace_font_line_by_line(self):
        book, _ = self.build("## C\n\n```python\nprint(1)\nprint(2)\n```\n")
        sheet = book["C"]
        rows = [row for row in self.values(sheet) if row != (None,)]
        self.assertEqual(rows, [("print(1)",), ("print(2)",)])
        self.assertEqual(sheet["A2"].font.name, "Consolas")

    def test_quotes_are_italic(self):
        book, _ = self.build("## Q\n\n> nhắc nhở\n")
        self.assertEqual(book["Q"]["A1"].value, "nhắc nhở")
        self.assertTrue(book["Q"]["A1"].font.italic)

    def test_front_matter_is_reported_not_written(self):
        book, warnings = self.build(
            "---\ntitle: X\n---\n\n## T\n\nnội dung\n"
        )
        self.assertEqual(self.values(book["T"]), [("nội dung",)])
        self.assertTrue(any("front matter" in w for w in warnings), warnings)


class DispatchTests(XlsxTestCase):
    def test_convert_file_writes_a_workbook_when_asked(self):
        source = self.dir / "bảng.md"
        source.write_text("## S\n\n| A |\n|---|\n| 1 |\n", encoding="utf-8")
        result = convert_file(
            source, self.dir / "out", ConversionOptions(markdown_target=".xlsx")
        )
        self.assertEqual(result.status, STATUS_SUCCESS, result.message)
        self.assertEqual(result.output.suffix, ".xlsx")
        self.assertEqual(load_workbook(result.output).sheetnames, ["S"])

    def test_word_is_still_the_default(self):
        source = self.dir / "bảng.md"
        source.write_text("# A\n\nx\n", encoding="utf-8")
        result = convert_file(source, self.dir / "out")
        self.assertEqual(result.output.suffix, ".docx")

    def test_an_unknown_target_falls_back_to_word(self):
        source = self.dir / "bảng.md"
        source.write_text("# A\n\nx\n", encoding="utf-8")
        result = convert_file(
            source, self.dir / "out", ConversionOptions(markdown_target=".rtf")
        )
        self.assertEqual(result.output.suffix, ".docx")

    def test_output_suffix_only_bends_for_markdown(self):
        self.assertEqual(output_suffix(Path("a.md"), ".xlsx"), ".xlsx")
        self.assertEqual(output_suffix(Path("a.md")), ".docx")
        self.assertEqual(output_suffix(Path("a.docx"), ".xlsx"), ".md")



class ImageTests(XlsxTestCase):
    def picture(self, name: str = "hình/a.png", size=(1600, 900)) -> Path:
        target = self.dir / name
        target.parent.mkdir(parents=True, exist_ok=True)
        Image.new("RGB", size, "red").save(target)
        return target

    def test_an_image_is_embedded_where_it_stands(self):
        self.picture()
        book, warnings = self.build("## T\n\n![](hình/a.png)\n")
        sheet = book["T"]
        self.assertEqual(warnings, [])
        self.assertEqual(len(sheet._images), 1, "the picture must be in the file")
        anchor = sheet._images[0].anchor._from
        self.assertEqual((anchor.row, anchor.col), (0, 0))

    def test_the_row_is_made_as_tall_as_the_picture(self):
        self.picture(size=(800, 600))
        book, _ = self.build("## T\n\n![](hình/a.png)\n")
        # 800x600 is scaled to fit 640x480, and a point is 0.75 of a pixel.
        self.assertEqual(book["T"].row_dimensions[1].height, 480 * 0.75)

    def test_alt_text_becomes_a_caption_under_the_picture(self):
        self.picture()
        book, _ = self.build("## T\n\n![Sơ đồ](hình/a.png)\n")
        self.assertEqual(book["T"]["A2"].value, "Sơ đồ")
        self.assertTrue(book["T"]["A2"].font.italic)

    def test_a_title_after_the_path_does_not_break_the_lookup(self):
        self.picture()
        book, warnings = self.build('## T\n\n![](hình/a.png "ghi chú")\n')
        self.assertEqual(warnings, [])
        self.assertEqual(len(book["T"]._images), 1)

    def test_text_around_an_image_keeps_its_order(self):
        self.picture()
        book, _ = self.build("## T\n\nTrước. ![](hình/a.png) Sau.\n")
        sheet = book["T"]
        # The picture sits on the row between the two lines of text.
        self.assertEqual(
            [row[0] for row in self.values(sheet)], ["Trước.", None, "Sau."]
        )
        self.assertEqual(sheet._images[0].anchor._from.row, 1)

    def test_a_missing_image_keeps_its_path_and_is_reported(self):
        book, warnings = self.build("## T\n\n![Sơ đồ](hình/thiếu.png)\n")
        self.assertEqual(book["T"]["A1"].value, "Sơ đồ (hình/thiếu.png)")
        self.assertEqual(book["T"]._images, [])
        self.assertTrue(any("Không tìm thấy ảnh" in w for w in warnings), warnings)

    def test_a_remote_image_is_not_downloaded(self):
        book, warnings = self.build("## T\n\n![Xa](https://x.dev/a.png)\n")
        self.assertEqual(book["T"]["A1"].value, "Xa (https://x.dev/a.png)")
        self.assertTrue(any("ảnh ở xa" in w for w in warnings), warnings)

    def test_embedding_can_be_turned_off(self):
        self.picture()
        book, _ = self.build("## T\n\n![Sơ đồ](hình/a.png)\n", embed_images=False)
        self.assertEqual(book["T"]["A1"].value, "Sơ đồ (hình/a.png)")
        self.assertEqual(book["T"]._images, [])

    def test_without_pillow_the_path_is_kept_instead(self):
        self.picture()

        def no_pillow(_path):
            raise ImportError("You must install Pillow to fetch image objects")

        original = md_to_xlsx.XlsxImage
        md_to_xlsx.XlsxImage = no_pillow
        try:
            book, warnings = self.build("## T\n\n![Sơ đồ](hình/a.png)\n")
        finally:
            md_to_xlsx.XlsxImage = original
        self.assertEqual(book["T"]["A1"].value, "Sơ đồ (hình/a.png)")
        self.assertTrue(any("Pillow" in w for w in warnings), warnings)

    def test_an_image_in_a_table_cell_stays_a_path(self):
        self.picture()
        book, _ = self.build("## T\n\n| A |\n|---|\n| ![](hình/a.png) |\n")
        self.assertEqual(book["T"]["A2"].value, "hình/a.png")
        self.assertEqual(book["T"]._images, [])


class CellSafetyTests(XlsxTestCase):
    LONG = "dài " * 60

    def test_long_text_wraps_instead_of_running_over(self):
        book, _ = self.build(f"## T\n\n{self.LONG}\n")
        cell = book["T"]["A1"]
        self.assertTrue(cell.alignment.wrap_text)
        self.assertEqual(cell.alignment.vertical, "top")

    def test_a_text_column_is_wide_enough_to_wrap_in(self):
        book, _ = self.build(f"## T\n\n{self.LONG}\n")
        self.assertEqual(book["T"].column_dimensions["A"].width, 90.0)

    def test_prose_is_held_to_the_table_width_when_they_share_a_column(self):
        book, _ = self.build(f"## T\n\n{self.LONG}\n\n| A |\n|---|\n| 1 |\n")
        self.assertEqual(book["T"].column_dimensions["A"].width, 60.0)

    def test_a_long_table_cell_wraps_too(self):
        book, _ = self.build(f"## T\n\n| A |\n|---|\n| {self.LONG} |\n")
        self.assertTrue(book["T"]["A2"].alignment.wrap_text)

    def test_control_characters_do_not_break_the_conversion(self):
        book, _ = self.build("## T\n\nco\x07 ky\x0b tự\n")
        self.assertEqual(book["T"]["A1"].value, "co ky\n tự")

    def test_text_that_starts_with_an_equals_sign_stays_text(self):
        book, _ = self.build("## T\n\n| A |\n|---|\n| =SUM(1,2) |\n")
        cell = book["T"]["A2"]
        self.assertEqual(cell.value, "=SUM(1,2)")
        self.assertEqual(cell.data_type, "s", "a formula would show an error")

    def test_a_cell_longer_than_excel_allows_is_cut(self):
        book, warnings = self.build(
            "## T\n\n" + "x" * (MAX_CELL_LENGTH + 100) + "\n"
        )
        self.assertEqual(len(book["T"]["A1"].value), MAX_CELL_LENGTH)
        self.assertTrue(any("cắt bớt" in w for w in warnings), warnings)


class FontTests(XlsxTestCase):
    """The font and size chosen for Word apply to a workbook as well."""

    SETTINGS = XlsxSettings(font="Arial", font_size=14)

    def test_body_cells_use_the_chosen_font(self):
        book, _ = self.build("## T\n\nĐoạn văn.\n", settings=self.SETTINGS)
        font = book["T"]["A1"].font
        self.assertEqual((font.name, font.size), ("Arial", 14.0))

    def test_table_cells_use_it_too(self):
        book, _ = self.build(
            "## T\n\n| A |\n|---|\n| 1 |\n", settings=self.SETTINGS
        )
        sheet = book["T"]
        self.assertEqual(sheet["A1"].font.name, "Arial")
        self.assertTrue(sheet["A1"].font.bold, "the header stays bold")
        self.assertEqual(sheet["A2"].font.size, 14.0)

    def test_a_heading_is_sized_above_the_body(self):
        book, _ = self.build(
            "# T\n\n## S\n\n### Con\n\nx\n", settings=self.SETTINGS
        )
        # 14 pt body, and a level-3 heading is three levels below the cap.
        self.assertEqual(book["S"]["A1"].font.size, 16.0)
        self.assertEqual(book["S"]["A2"].font.size, 14.0)

    def test_code_keeps_consolas_at_the_chosen_size(self):
        book, _ = self.build(
            "## T\n\n```\nprint(1)\n```\n", settings=self.SETTINGS
        )
        font = book["T"]["A1"].font
        self.assertEqual((font.name, font.size), ("Consolas", 14.0))

    def test_a_caption_is_a_point_smaller(self):
        target = self.dir / "a.png"
        Image.new("RGB", (100, 100), "red").save(target)
        book, _ = self.build("## T\n\n![Chú thích](a.png)\n", settings=self.SETTINGS)
        self.assertEqual(book["T"]["A2"].font.size, 13.0)

    def test_a_link_keeps_the_font_and_gains_the_link_look(self):
        book, _ = self.build(
            "## T\n\n| A |\n|---|\n| [đi](https://x.dev) |\n",
            settings=self.SETTINGS,
        )
        font = book["T"]["A2"].font
        self.assertEqual((font.name, font.size), ("Arial", 14.0))
        self.assertEqual(font.underline, "single")

    def test_nothing_chosen_leaves_excels_own_font_alone(self):
        book, _ = self.build("## T\n\nĐoạn.\n")
        font = book["T"]["A1"].font
        self.assertIsNone(font.name)
        self.assertIsNone(font.size)

if __name__ == "__main__":
    unittest.main(verbosity=2)
