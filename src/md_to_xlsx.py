"""Markdown -> Excel (.xlsx).

The mirror image of `excel_to_markdown`: every top-level section of the document
becomes one worksheet, and every Markdown table inside it becomes a real grid of
cells. Anything that is not a table - headings, paragraphs, lists, quotes, code -
is written down column A in document order, so nothing from the source is lost.
Images are embedded as pictures; one that cannot be embedded keeps its path as
text rather than taking the whole conversion down with it.
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import unquote, urlparse

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlsxImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .md_to_docx import (
    CodeBlock,
    Heading,
    ListBlock,
    ListItem,
    Paragraph,
    Quote,
    Rule,
    Span,
    Table,
    parse_inline,
    parse_markdown,
)

# Excel's own limits, not ours.
MAX_SHEET_NAME = 31
MAX_CELL_LENGTH = 32767
_INVALID_SHEET_CHARS = re.compile(r"[\[\]:*?/\\]")

# openpyxl rejects these outright. Word leaves \x0b behind for a soft line break
# and \x0c for a page break, so those two become real newlines instead.
_ILLEGAL_CHARS = re.compile(r"[\x00-\x08\x0e-\x1f]")
_BREAK_CHARS = re.compile(r"[\x0b\x0c]")

# Only a plain integer or decimal becomes a number. "007", "1,5" and "12%" stay
# text: turning them into numbers would change what the document says.
_NUMBER_RE = re.compile(r"^-?(?:0|[1-9][0-9]{0,14})(?:\.[0-9]+)?$")

MIN_COLUMN_WIDTH = 8.0
MAX_COLUMN_WIDTH = 60.0
# Column A carries prose, and gets more room when no table has claimed it.
MAX_TEXT_COLUMN_WIDTH = 90.0

# A picture is scaled down to fit this box, in pixels.
MAX_IMAGE_WIDTH = 640
MAX_IMAGE_HEIGHT = 480
_POINTS_PER_PIXEL = 0.75
_DEFAULT_ROW_HEIGHT = 15.0

_REMOTE_SCHEMES = {"http", "https", "data", "ftp"}

DEFAULT_SHEET_NAME = "Sheet1"

_HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
_CODE_FONT = "Consolas"
# What Excel falls back to, and what a heading is sized against.
_EXCEL_FONT_SIZE = 11.0
_BULLETS = ["•", "◦", "‣"]


@dataclass(frozen=True)
class XlsxSettings:
    """How the generated workbook should look.

    An empty `font` or a `font_size` of zero leaves Excel's own default alone.
    """

    font: str = ""
    font_size: float = 0.0
    freeze_header: bool = True
    numbers_as_numbers: bool = True
    wrap_text: bool = True

    @property
    def base_size(self) -> float:
        """The size headings and captions are measured against."""
        return self.font_size or _EXCEL_FONT_SIZE


def markdown_to_xlsx(
    source: str | Path,
    destination: str | Path,
    embed_images: bool = True,
    add_title_heading: bool = True,
    settings: XlsxSettings | None = None,
) -> list[str]:
    """Write `source` as a workbook at `destination`, returning any warnings.

    `add_title_heading` decides where a nameless sheet gets its name from: the
    file name when it is on, a plain "Sheet1" when it is off.
    """
    settings = settings or XlsxSettings()
    source = Path(source)
    destination = Path(destination)
    warnings: list[str] = []

    text = source.read_text(encoding="utf-8-sig", errors="replace")
    blocks, front_matter = parse_markdown(text)
    if front_matter:
        warnings.append(
            f"Bỏ qua front matter YAML ({len(front_matter)} khoá) - Excel không có "
            "chỗ cho phần này."
        )

    # A worksheet must be called something: sections name themselves, and what
    # is left over borrows the file name unless the user asked us not to.
    fallback = source.stem if add_title_heading else DEFAULT_SHEET_NAME
    sheets = _split_sheets(blocks, fallback)
    workbook = Workbook()
    workbook.remove(workbook.active)

    if not sheets:
        warnings.append("Tài liệu không có nội dung; workbook chỉ có một sheet trống.")
        sheets = [(fallback, [])]

    used: set[str] = set()
    for name, section in sheets:
        sheet = workbook.create_sheet(_sheet_name(name or fallback, used))
        _Writer(
            sheet,
            settings,
            base_dir=source.parent,
            embed_images=embed_images,
            warnings=warnings,
        ).write(section)

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return warnings


# ------------------------------------------------------------------ sheets


def _sheet_level(blocks) -> int:
    """The heading level that starts a new worksheet.

    Usually the shallowest heading in the document. A single `#` title with `##`
    sections below it is the shape `excel_to_markdown` produces, so a lone top
    heading steps aside and the level below it splits the sheets instead.
    """
    levels = sorted({block.level for block in blocks if isinstance(block, Heading)})
    if not levels:
        return 0
    top = levels[0]
    tops = sum(1 for b in blocks if isinstance(b, Heading) and b.level == top)
    if tops == 1 and len(levels) > 1:
        return levels[1]
    return top


def _split_sheets(blocks, fallback: str) -> list[tuple[str, list]]:
    """Cut the document into (sheet name, blocks) pairs."""
    level = _sheet_level(blocks)
    sheets: list[tuple[str, list]] = []
    name = fallback
    current: list = []

    for block in blocks:
        if level and isinstance(block, Heading) and block.level <= level:
            if current:
                sheets.append((name, current))
            name, current = _plain_text(block.text) or fallback, []
            continue
        current.append(block)

    if current:
        sheets.append((name, current))
    return sheets


def _sheet_name(name: str, used: set[str]) -> str:
    """A name Excel accepts, and one it has not seen in this workbook yet."""
    cleaned = _INVALID_SHEET_CHARS.sub("", name).strip().strip("'")
    cleaned = re.sub(r"\s+", " ", cleaned)[:MAX_SHEET_NAME] or DEFAULT_SHEET_NAME

    candidate = cleaned
    counter = 2
    while candidate.casefold() in used:
        tail = f" ({counter})"
        candidate = cleaned[: MAX_SHEET_NAME - len(tail)].strip() + tail
        counter += 1
    used.add(candidate.casefold())
    return candidate


# -------------------------------------------------------------------- cells


def _clean(text: str) -> str:
    """Text a cell will actually accept: openpyxl rejects control characters."""
    return _ILLEGAL_CHARS.sub("", _BREAK_CHARS.sub("\n", text))


def _image_text(span: Span) -> str:
    alt = span.text.strip()
    return f"{alt} ({span.src})" if alt else span.src


def _spans_text(spans) -> tuple[str, str]:
    """What a run of spans reads as, plus the URL all of it links to, if any.

    A cell that is nothing but a link becomes a real Excel hyperlink. When a
    link only makes up part of the text its target is spelled out in brackets
    instead, so no URL is lost on the way into the spreadsheet.
    """
    linked = {span.style.href for span in spans if span.style.href}
    whole = len(linked) == 1 and all(
        span.style.href or not span.text.strip() for span in spans
    )

    pieces: list[str] = []
    previous = ""

    def close(href: str) -> None:
        if href and not whole and href not in "".join(pieces):
            pieces.append(f" ({href})")

    for span in spans:
        if span.style.href != previous:
            close(previous)
        previous = span.style.href
        if span.kind == "break":
            pieces.append("\n")
        elif span.kind == "image":
            pieces.append(_image_text(span))
        else:
            pieces.append(span.text)
    close(previous)

    return "".join(pieces).strip(), (linked.pop() if whole and linked else "")


def _plain_text(text: str) -> str:
    """Inline Markdown as the plain text a sheet name or heading can hold."""
    return _spans_text(parse_inline(text))[0]


def _cell_text(text: str) -> tuple[str, str]:
    """Inline Markdown as (cell text, hyperlink target)."""
    return _spans_text(parse_inline(text))


def _segments(text: str) -> list[tuple[str, object]]:
    """Split a paragraph into runs of text and the images between them.

    A picture cannot live inside a cell, so a standalone image goes onto the
    sheet on its own while the text around it stays in column A.
    """
    result: list[tuple[str, object]] = []
    run: list[Span] = []
    for span in parse_inline(text):
        if span.kind == "image":
            if run:
                result.append(("text", run))
                run = []
            result.append(("image", span))
        else:
            run.append(span)
    if run:
        result.append(("text", run))
    return result


def _strip_title(src: str) -> str:
    """Drop the optional "title" after an image path, and its angle brackets."""
    src = src.strip()
    for quote_char in ('"', "'"):
        marker = f" {quote_char}"
        if src.endswith(quote_char) and marker in src:
            src = src[: src.index(marker)]
    return src.strip().lstrip("<").rstrip(">").strip()


def _scaled(width: int, height: int) -> tuple[int, int]:
    """Shrink a picture to fit the box above, keeping its proportions."""
    if not width or not height:
        return MAX_IMAGE_WIDTH, MAX_IMAGE_HEIGHT
    ratio = min(1.0, MAX_IMAGE_WIDTH / width, MAX_IMAGE_HEIGHT / height)
    return max(1, int(width * ratio)), max(1, int(height * ratio))


def _cell_value(text: str, numbers: bool):
    """A number when the text is plainly one, otherwise the text itself."""
    if not numbers or not _NUMBER_RE.match(text):
        return text
    return float(text) if "." in text else int(text)


def _align(alignments: list[str], column: int) -> str:
    return alignments[column - 1] if column <= len(alignments) else ""


def _apply_link(cell, href: str) -> None:
    """Turn a cell into a clickable link without losing the font it had."""
    if not href:
        return
    cell.hyperlink = href
    base = cell.font
    cell.font = Font(
        name=base.name,
        size=base.size,
        bold=base.bold,
        italic=base.italic,
        color="0563C1",
        underline="single",
    )


class _Writer:
    """Fills one worksheet, keeping track of rows, widths and warnings."""

    def __init__(
        self,
        sheet,
        settings: XlsxSettings,
        base_dir: Path,
        embed_images: bool,
        warnings: list[str],
    ):
        self.sheet = sheet
        self.settings = settings
        self.base_dir = base_dir
        self.embed_images = embed_images
        self.warnings = warnings
        self.row = 1
        self.widths: dict[int, float] = {}
        self.text_widths: dict[int, float] = {}
        self.first_block = True
        self.froze = False

    # ------------------------------------------------------------- helpers

    def warn(self, message: str) -> None:
        if message not in self.warnings:  # the same complaint once is enough
            self.warnings.append(message)

    def note_width(self, column: int, text: str, table: bool = True) -> None:
        longest = max((len(line) for line in str(text).split("\n")), default=0)
        cap = MAX_COLUMN_WIDTH if table else MAX_TEXT_COLUMN_WIDTH
        width = min(max(longest + 2, MIN_COLUMN_WIDTH), cap)
        target = self.widths if table else self.text_widths
        target[column] = max(target.get(column, 0.0), width)

    def put(
        self,
        column: int,
        value,
        font: Font | None = None,
        alignment: Alignment | None = None,
        href: str = "",
    ):
        """Write one cell, keeping Excel out of trouble."""
        cell = self.sheet.cell(row=self.row, column=column)
        if isinstance(value, str):
            value = _clean(value)
            if len(value) > MAX_CELL_LENGTH:
                value = value[:MAX_CELL_LENGTH]
                self.warn(f"Có ô dài hơn {MAX_CELL_LENGTH} ký tự nên bị cắt bớt.")
        cell.value = value
        # A cell that happens to start with "=" is text here, not a formula.
        if isinstance(value, str) and value.startswith("="):
            cell.data_type = "s"
        if font is not None:
            cell.font = font
        if alignment is not None:
            cell.alignment = alignment
        _apply_link(cell, href)
        return cell

    def font(
        self,
        bold: bool = False,
        italic: bool = False,
        size: float = 0.0,
        name: str = "",
    ) -> Font:
        """The chosen font, with one attribute or another turned up.

        Every cell is given one explicitly: a cell with no font of its own
        falls back to Calibri 11 when the workbook is reopened.
        """
        return Font(
            name=name or self.settings.font or None,
            size=size or self.settings.font_size or None,
            bold=bold or None,
            italic=italic or None,
        )

    def alignment(self, horizontal: str = "", indent: int = 0) -> Alignment:
        """Wrapped and top-aligned, so a long cell grows down instead of over."""
        return Alignment(
            horizontal=horizontal or None,
            vertical="top",
            wrap_text=self.settings.wrap_text,
            indent=indent,
        )

    def label(
        self, text: str, font: Font | None = None, indent: int = 0, href: str = ""
    ) -> None:
        """One line of text in column A."""
        self.put(
            1,
            text,
            font=font or self.font(),
            alignment=self.alignment(indent=indent),
            href=href,
        )
        self.note_width(1, text, table=False)
        self.row += 1

    def blank(self) -> None:
        if self.row > 1:
            self.row += 1

    def finish(self) -> None:
        # A column is as wide as the widest thing in it, but prose is held to
        # the narrower table cap wherever a table shares the column.
        for column in set(self.widths) | set(self.text_widths):
            table = self.widths.get(column, 0.0)
            prose = self.text_widths.get(column, 0.0)
            if table:
                prose = min(prose, MAX_COLUMN_WIDTH)
            letter = get_column_letter(column)
            self.sheet.column_dimensions[letter].width = max(table, prose)

    # -------------------------------------------------------------- blocks

    def write(self, blocks) -> None:
        for block in blocks:
            self.block(block)
            self.first_block = False
        self.finish()

    def block(self, block) -> None:
        if isinstance(block, Table):
            self.table(block)
        elif isinstance(block, Heading):
            self.heading(block)
        elif isinstance(block, Paragraph):
            self.paragraph(block.text)
        elif isinstance(block, ListBlock):
            self.list_block(block)
        elif isinstance(block, CodeBlock):
            self.code(block)
        elif isinstance(block, Quote):
            self.quote(block)
        elif isinstance(block, Rule):
            self.blank()

    def heading(self, block: Heading) -> None:
        text = _plain_text(block.text)
        if not text:
            return
        self.blank()
        # A heading stands out from the body text it was set against.
        size = self.settings.base_size + max(1, 5 - block.level)
        self.label(text, self.font(bold=True, size=size))

    def paragraph(self, text: str, font: Font | None = None, indent: int = 0) -> None:
        for kind, payload in _segments(text):
            if kind == "image":
                self.image(payload)
                continue
            plain, href = _spans_text(payload)
            if not plain:
                continue
            for line in plain.split("\n"):
                self.label(line, font, indent=indent, href=href)

    def quote(self, block: Quote) -> None:
        for inner in block.blocks:
            if isinstance(inner, Paragraph):
                self.paragraph(inner.text, self.font(italic=True), indent=1)
            else:
                self.block(inner)

    def code(self, block: CodeBlock) -> None:
        self.blank()
        for line in block.text.split("\n"):
            self.label(line, self.font(name=_CODE_FONT))

    def list_block(self, block: ListBlock, depth: int = 0) -> None:
        number = block.start
        for item in block.items:
            self.list_item(block, item, number, depth)
            number += 1

    def list_item(self, block: ListBlock, item: ListItem, number: int, depth: int):
        marker = f"{number}." if block.ordered else _BULLETS[depth % len(_BULLETS)]
        if item.checked is not None:
            marker += " [x]" if item.checked else " [ ]"

        first = True
        for inner in item.blocks:
            if isinstance(inner, ListBlock):
                self.list_block(inner, depth + 1)
                continue
            if not isinstance(inner, (Paragraph, CodeBlock)):
                self.block(inner)
                continue
            plain, href = _cell_text(inner.text)
            for line in plain.split("\n"):
                if not line and first:
                    continue  # an item with no text would leave a lone bullet
                prefix = f"{marker} " if first else ""
                self.label(f"{prefix}{line}", indent=depth + 1, href=href)
                first = False

    def table(self, block: Table) -> None:
        self.blank()
        header_row = self.row
        alignments = list(block.alignments)

        for column, text in enumerate(block.header, start=1):
            plain, href = _cell_text(text)
            cell = self.put(
                column,
                plain,
                font=self.font(bold=True),
                alignment=self.alignment(_align(alignments, column)),
                href=href,
            )
            cell.fill = _HEADER_FILL
            self.note_width(column, plain)
        self.row += 1

        for row_cells in block.rows:
            for column, text in enumerate(row_cells, start=1):
                plain, href = _cell_text(text)
                self.put(
                    column,
                    _cell_value(plain, self.settings.numbers_as_numbers),
                    font=self.font(),
                    alignment=self.alignment(_align(alignments, column)),
                    href=href,
                )
                self.note_width(column, plain)
            self.row += 1

        # Only a sheet that opens with its table gets frozen headers: a pane
        # split further down the sheet would hide the rows above it.
        if self.settings.freeze_header and self.first_block and not self.froze:
            self.sheet.freeze_panes = f"A{header_row + 1}"
            self.froze = True

    # -------------------------------------------------------------- images

    def image(self, span: Span) -> None:
        """Put a picture on the sheet, or its path when that cannot be done."""
        path = self._resolve(span.src)
        if path is None:
            fallback = _image_text(span)
            if fallback:
                self.label(fallback)
            return

        try:
            picture = XlsxImage(str(path))
        except ImportError:  # Pillow is how openpyxl reads an image at all
            self.warn(
                "Chưa cài Pillow nên không nhúng được ảnh vào Excel; "
                "chỉ giữ lại đường dẫn."
            )
            self.label(_image_text(span))
            return
        except Exception as exc:  # noqa: BLE001 - one bad image, not one bad file
            self.warn(f"Không chèn được ảnh {span.src}: {exc}")
            self.label(_image_text(span))
            return

        picture.width, picture.height = _scaled(picture.width, picture.height)
        self.sheet.add_image(picture, f"A{self.row}")
        # The row is made as tall as the picture, so nothing below is covered.
        self.sheet.row_dimensions[self.row].height = max(
            _DEFAULT_ROW_HEIGHT, picture.height * _POINTS_PER_PIXEL
        )
        self.row += 1

        caption = span.text.strip()
        if caption:
            size = max(8.0, self.settings.base_size - 1)
            self.label(caption, self.font(italic=True, size=size))

    def _resolve(self, src: str) -> Path | None:
        """The file an image points at, or None when it cannot be embedded."""
        if not self.embed_images or not src:
            return None
        target = _strip_title(src)
        if urlparse(target).scheme in _REMOTE_SCHEMES:
            self.warn(f"Bỏ qua ảnh ở xa (không tải về): {target}")
            return None
        candidate = Path(unquote(target))
        if not candidate.is_absolute():
            candidate = self.base_dir / candidate
        if candidate.is_file():
            return candidate
        self.warn(f"Không tìm thấy ảnh: {src}")
        return None
